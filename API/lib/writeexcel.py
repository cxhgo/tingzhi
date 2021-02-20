# _*_ coding:utf-8 _*_
import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
from API.config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
#from openpyxl.styles.colors import RED,GREEN,DARKYELLOW
from openpyxl.styles import colors
import configparser as cparser
from jsonpath_rw import parse
# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.Test_Config,encoding='UTF-8')
name = cf.get("tester","name")

class WriteExcel():
    # -------------文件写入数据---------------
    def __init__(self,filename,sheetname):
        """
        初始化数据
        :param filename:xls文件名
        :param sheetname:表格名称
        :return: 无
        """
        self.filename = filename
        # 如果不存在文件，则将文件1覆盖到文件2
        if not os.path.exists(self.filename):
            shutil.copyfile(setting.Source_File,setting.Target_File)
        # 打开文件表格，获取文件表
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[sheetname]


    def write_data(self,row_num,value):
        """
        将测试结果和测试人员写入表格
        :param row_num:对应行数
        :param value:测试结果值
        :return: 无
        """
        # 写入测试结果，定义颜色、显示
        font_Creen = Font(name='宋体',color= "0000FF00",bold=True)
        font_Red = Font(name='宋体',color="00FFBB00",bold=True)
        font_Yellow = Font(name='宋体',color="00FFFF00",bold=True)
        align = Alignment(horizontal='center',vertical='center')
        # font_Creen = Font(name='宋体',color= GREEN,bold=True)
        # font_Red = Font(name='宋体',color=RED,bold=True)
        # font_Yellow = Font(name='宋体',color=DARKYELLOW,bold=True)
        # 获取写入结果的所在行数
        L_num = "L"+str(row_num)
        M_num = "M"+str(row_num)
        # 在第row_n行第12列写入value
        if value =="PASS":
            self.ws.cell(row_num,12,value)
            self.ws[L_num].font = font_Creen
        if value =="FAIL":
            self.ws.cell(row_num,12,value)
            self.ws[L_num].font = font_Red
        # 在第row_num行第13列写入测试人员名字
        self.ws.cell(row_num,13,name)
        self.ws[L_num].alignment =align
        self.ws[M_num].alignment = align
        self.ws[M_num].font = font_Yellow
        self.wb.save(self.filename)

    def connectdata_write(self,take_key,result,rowNum):
        """
        找到关联接口的关键字，取出给关联用例作为参数传递
        :param take_key:需要取出的关键字
        :param result:接口返回的响应内容
        :param rowNum:对应需要取出数据的接口的行数，用于写入取出的数据保存到表格
        :return: 无
        """
        if take_key!="":

             print("表格中需要从响应数据中取出的关键字take_key：",take_key)
             #有多个的关键字
             if "," in take_key:
               take_list = take_key.split(",")
               print("多个需要从响应数据中取出的关键字以逗号划分重新组合的列表take_list：",take_list)
               key_list=[]
               for i in take_list:
                   print("遍历的关键字列表中的关键字：",i)
                   try:
                     test= [match.value for match in parse(str(i)).find(result)][0]
                   except Exception:
                     print("从返回的响应数据中找到目标-关联的下一个接口的数据：",test)
                   if test =="":
                      print("获取的对应目标数据为空!")
                   else:
                      key_list.append(test)
                      print("下一个接口的关键字列表key_list:",key_list)
                      self.ws.cell(rowNum,17,str(key_list))
                      self.wb.save(self.filename)
            #只有单个的关键字
             else:
                re_content = [match.value for match in parse(str(take_key)).find(result)][0]
                print("找到对应响应数据的内容从返回的响应数据中找到目标-关联的下一个接口的数据：",re_content)
                self.ws.cell(rowNum,17,str(re_content))
                self.wb.save(self.filename)



    def clear_write(self):
        """
        清除当天执行的从响应内容获取的关联数据
        :param ：无
        :return: 无
        """
        clear_large=38
        for i in range(2,clear_large):
            self.ws.cell(i,17,value="")
        self.wb.save(self.filename)
        print("清除成功！")